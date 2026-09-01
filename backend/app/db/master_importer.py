"""積算コードMaster Importer (Phase 1.7、および使用品名限定の追加指示分)。

処理の流れ (指示書6章 / 追加指示7章):

    data/master/estimate_master_a.xlsx
            │
            ▼
    Sheet2
            │
            ▼
    行読込
            │
            ▼
    コード/品名の取り消し線判定 (無効なら除外)
            │
            ▼
    対象13品名か判定 (`app.domain.master_categories.ALLOWED_CATEGORIES` にあるか)
            │
            ▼
    有効行のみ正規化してUPSERT
            │
            ▼
    SQLite estimate_master_items
            │
            ▼
    REST API (/api/master-items)
            │
            ▼
    Frontend

Frontendから直接Excelを読むことはせず、Excelの列構造・シート名・表記揺れ・
取り消し線判定・対象品名の絞り込みといった詳細は全てこのモジュールに閉じ込める
(Frontendは正規化・フィルタ済みのDomain Modelのみを見る)。

シート構成の調査結果 (read-only で確認済み):
  - "Sheet2" (912データ行) が対象品目の全件データ。列は
    コード/品名/型式/定格/総合価格A/箱・部品価格/塗装価格/設A/板金/組立/検査 の11列
    (12列目は空欄のことが多いが、一部の行では備考的なテキストが入っている)。
  - "Sheet1" は同一データの一部を書き写した作業用シート (147行のみ、以降の列に
    無関係なメモが混在) であり、Masterの正式データとしては使用しない。
  - コード列は912件全てユニークであることを確認済み (unique keyとして採用)。

使用品名の限定 (追加指示1章/2章):
  Sekisan Naviで使用するのは `master_categories.ALLOWED_CATEGORIES` の13品名のみ。
  それ以外のcategory (文章形式の特殊行4件、品名空欄1件を含む) は取り込まない。
  「未分類」タブでの表示も廃止した (該当行はそもそもDBへ入れない)。

取り消し線の判定 (追加指示4章/5章):
  openpyxlでセル書式(`cell.font.strike`)を取得できるよう、`values_only=False`で
  行を読み込む。コードセルまたは品名セルのどちらかに取り消し線(strike=True)が
  設定されている行は無効行として取り込まない。文字列パターンによる推測は行わない。

同期方式 (再取込, 指示書7章 / 追加指示8章/9章):
  コード(code)をunique keyとしたUPSERTで取り込む。
    - 新規コード: 新規行としてINSERT (新しいidが振られる)
    - 既存コード: 既存行のid を維持したまま値のみUPDATE
  今回の条件 (13品名限定・取り消し線除外) で無効となった既存Master行のうち、
  どのDetectionからも参照されていないものは安全のため削除する
  (Excel側で対象外になったデータをDBに残さない)。
  ただし既存のManual BBox (detections.master_item_id) が参照している行は、
  無効化条件に該当していても勝手に削除しない (FK違反の回避、ユーザーデータ保護)。
  該当があった場合は `MasterImportResult.retained_invalid_referenced` へ記録し、
  呼び出し側で報告できるようにする。
"""
from __future__ import annotations

import sqlite3
from dataclasses import dataclass, field
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell

from app.config import MASTER_EXCEL_PATH, MASTER_EXCEL_SHEET
from app.domain.master_categories import ALLOWED_CATEGORIES

_ALLOWED_CATEGORY_SET = set(ALLOWED_CATEGORIES)


class MasterImportError(Exception):
    """Excelファイルが見つからない・想定した構成でない場合に送出する。"""


@dataclass
class MasterImportResult:
    total_rows_in_sheet: int
    imported: int
    inserted: int
    updated: int
    skipped_no_code: int
    excluded_by_strike: int
    excluded_by_category: int
    # (code, category) のペア。取り消し線により除外した行のコード一覧 (完了報告用)。
    excluded_by_strike_codes: list[str] = field(default_factory=list)
    # 今回の条件で無効となり、DBから削除した既存Master行の件数。
    removed_stale: int = 0
    # 今回の条件で無効になったが、既存Manual BBoxが参照しているため削除せず残した
    # Master行の (id, code) 一覧。指示書9章の通り、ユーザーデータを壊さないための記録。
    retained_invalid_referenced: list[tuple[int, str]] = field(default_factory=list)


def _to_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def _is_struck(cell: Cell | None) -> bool:
    """セルのフォントに取り消し線が設定されているかどうかを判定する。

    文字列の内容から推測するのではなく、Excelの実際のセル書式 (`font.strike`) を見る
    (追加指示5章)。
    """
    if cell is None or cell.font is None:
        return False
    return bool(cell.font.strike)


def import_master_excel(
    conn: sqlite3.Connection,
    excel_path: Path = MASTER_EXCEL_PATH,
    sheet_name: str = MASTER_EXCEL_SHEET,
) -> MasterImportResult:
    """積算コードMaster Excelを読み込み、estimate_master_itemsへ全件UPSERTする。

    件数制限・先頭N件だけの取込・NULLの補完は一切行わない。ただし追加指示により、
    取り消し線が設定された行、および対象13品名以外の行は取り込み対象から除外する。
    """
    if not excel_path.exists():
        raise MasterImportError(f"Master Excelファイルが見つかりません: {excel_path}")

    # セル書式(取り消し線)を判定する必要があるため、values_only=False (既定) で読み込む。
    # data_only=True は数式セルの計算済みキャッシュ値を使う指定であり、フォント情報の
    # 取得可否には影響しない。
    workbook = openpyxl.load_workbook(excel_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise MasterImportError(
            f"Master Excelに想定シート '{sheet_name}' が見つかりません "
            f"(実際のシート: {workbook.sheetnames})"
        )
    worksheet = workbook[sheet_name]

    total_rows = 0
    inserted = 0
    updated = 0
    skipped_no_code = 0
    excluded_by_strike = 0
    excluded_by_category = 0
    excluded_by_strike_codes: list[str] = []
    valid_codes: set[str] = set()

    for row in worksheet.iter_rows(min_row=2):
        values = [c.value for c in row]
        if all(v is None for v in values):
            continue
        total_rows += 1

        code_cell = row[0] if len(row) > 0 else None
        category_cell = row[1] if len(row) > 1 else None

        code_raw = code_cell.value if code_cell is not None else None
        if code_raw is None:
            skipped_no_code += 1
            continue
        code = str(code_raw).strip()
        if not code:
            skipped_no_code += 1
            continue

        # 処理概念 (追加指示7章) の通り、取り消し線判定 -> 対象品名判定 の順で行う。
        if _is_struck(code_cell) or _is_struck(category_cell):
            excluded_by_strike += 1
            excluded_by_strike_codes.append(code)
            continue

        category = _to_text(category_cell.value if category_cell is not None else None)
        if category not in _ALLOWED_CATEGORY_SET:
            excluded_by_category += 1
            continue

        model = _to_text(row[2].value) if len(row) > 2 else None
        rating = _to_text(row[3].value) if len(row) > 3 else None
        total_price_a = row[4].value if len(row) > 4 else None
        box_parts_price = row[5].value if len(row) > 5 else None
        painting_price = row[6].value if len(row) > 6 else None
        setup_a = row[7].value if len(row) > 7 else None
        sheet_metal_price = row[8].value if len(row) > 8 else None
        assembly_price = row[9].value if len(row) > 9 else None
        inspection_price = row[10].value if len(row) > 10 else None
        note = _to_text(row[11].value) if len(row) > 11 else None

        existing = conn.execute(
            "SELECT id FROM estimate_master_items WHERE code = ?", (code,)
        ).fetchone()

        conn.execute(
            """
            INSERT INTO estimate_master_items
                (code, category, model, rating, note, total_price_a, box_parts_price,
                 painting_price, setup_a, sheet_metal_price, assembly_price, inspection_price)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(code) DO UPDATE SET
                category = excluded.category,
                model = excluded.model,
                rating = excluded.rating,
                note = excluded.note,
                total_price_a = excluded.total_price_a,
                box_parts_price = excluded.box_parts_price,
                painting_price = excluded.painting_price,
                setup_a = excluded.setup_a,
                sheet_metal_price = excluded.sheet_metal_price,
                assembly_price = excluded.assembly_price,
                inspection_price = excluded.inspection_price
            """,
            (
                code, category, model, rating, note, total_price_a, box_parts_price,
                painting_price, setup_a, sheet_metal_price, assembly_price, inspection_price,
            ),
        )
        valid_codes.add(code)
        if existing:
            updated += 1
        else:
            inserted += 1

    removed_stale, retained_invalid_referenced = _sync_remove_stale_master_items(
        conn, valid_codes
    )

    return MasterImportResult(
        total_rows_in_sheet=total_rows,
        imported=inserted + updated,
        inserted=inserted,
        updated=updated,
        skipped_no_code=skipped_no_code,
        excluded_by_strike=excluded_by_strike,
        excluded_by_category=excluded_by_category,
        excluded_by_strike_codes=excluded_by_strike_codes,
        removed_stale=removed_stale,
        retained_invalid_referenced=retained_invalid_referenced,
    )


def _sync_remove_stale_master_items(
    conn: sqlite3.Connection, valid_codes: set[str]
) -> tuple[int, list[tuple[int, str]]]:
    """今回の条件で無効となった既存Master行を、安全な範囲で削除する (追加指示8章/9章)。

    「今回の条件で無効」= 今回のExcel取込パスで `valid_codes` に含まれなかった
    既存行 (=取り消し線・対象外品名等により除外された行、またはExcelから
    削除されたコード)。

    単純な全DELETE→再INSERTは行わない。既存のManual BBox (detections.master_item_id)
    が参照している行は、無効化条件に該当していても削除せず、
    `retained_invalid_referenced` として呼び出し側へ報告する。
    """
    referenced_ids = {
        row["master_item_id"]
        for row in conn.execute(
            "SELECT DISTINCT master_item_id FROM detections WHERE master_item_id IS NOT NULL"
        ).fetchall()
    }

    stale_rows = conn.execute(
        "SELECT id, code FROM estimate_master_items"
    ).fetchall()

    removed = 0
    retained: list[tuple[int, str]] = []
    for row in stale_rows:
        if row["code"] in valid_codes:
            continue
        if row["id"] in referenced_ids:
            retained.append((row["id"], row["code"]))
            continue
        conn.execute("DELETE FROM estimate_master_items WHERE id = ?", (row["id"],))
        removed += 1

    return removed, retained


def main() -> None:
    """再取込用のスタンドアロン実行エントリポイント。

    使い方: python -m app.db.master_importer
    """
    from app.config import DB_PATH
    from app.db.connection import get_connection
    from app.db.migrate import run_migrations

    run_migrations(DB_PATH)
    with get_connection(DB_PATH) as conn:
        result = import_master_excel(conn)
    print(
        f"Master import completed: imported={result.imported} "
        f"(inserted={result.inserted}, updated={result.updated}), "
        f"skipped_no_code={result.skipped_no_code}, "
        f"excluded_by_strike={result.excluded_by_strike}, "
        f"excluded_by_category={result.excluded_by_category}, "
        f"removed_stale={result.removed_stale}, "
        f"retained_invalid_referenced={result.retained_invalid_referenced}, "
        f"total_rows_in_sheet={result.total_rows_in_sheet}"
    )


if __name__ == "__main__":
    main()
