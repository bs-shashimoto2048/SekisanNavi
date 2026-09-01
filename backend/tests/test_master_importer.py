"""積算コードMaster Importer (Phase 1.7、および使用品名限定の追加指示) のテスト。

実際にプロジェクトへ同梱されている `data/master/estimate_master_a.xlsx` は
アプリが参照する安定した正式データであるため、そのままテスト対象として使う
(全対象行の取込・使用13品名への絞り込み・10列マッピング・NULL保持を検証する)。
あわせて、UPSERT(再取込)・取り消し線判定・対象外品名の除外・安全な削除同期の
細かな挙動は openpyxl でその場に作った小さなExcelを使って検証する。

品名の文字列は半角/全角の表記ゆれ (例: 半角中点｢箱･単独｣ vs 全角｢箱・単独｣) を
手で打ち間違えないよう、テストコード中でも `app.domain.master_categories` から
インポートした定数をそのまま使う (要件: 二重管理・転記ミスの防止)。
"""
import sqlite3

import openpyxl
import pytest
from openpyxl.styles import Font

from app.config import MASTER_EXCEL_PATH, MASTER_EXCEL_SHEET
from app.db.connection import get_connection
from app.db.master_importer import MasterImportError, import_master_excel
from app.db.migrate import run_migrations
from app.domain.master_categories import ALLOWED_CATEGORIES


# --- 実Excelに対するテスト (対象13品名への絞り込み・取り消し線除外・NULL保持の確認) ---


def test_real_master_excel_imports_all_valid_rows(client, db_path):
    """Sheet2の912行のうち、取り消し線(3件)・対象外品名(4件)を除いた905件のみ取り込む。"""
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    count = conn.execute("SELECT COUNT(*) AS c FROM estimate_master_items").fetchone()["c"]
    conn.close()
    assert count == 905


def test_real_master_excel_category_counts_match_source(client):
    res = client.get("/api/master-items", params={"category": ALLOWED_CATEGORIES[0]})  # 箱･単独
    assert len(res.json()) == 230

    res = client.get("/api/master-items", params={"category": ALLOWED_CATEGORIES[1]})  # 箱･左右
    assert len(res.json()) == 230

    res = client.get("/api/master-items", params={"category": ALLOWED_CATEGORIES[6]})  # 附属品加算価格
    assert len(res.json()) == 29

    # 箱体価格倍率は元は21件だが、うち2件(コード19958/19960)が取り消し線のため
    # 19件のみ取り込まれる。
    res = client.get("/api/master-items", params={"category": ALLOWED_CATEGORIES[7]})  # 箱体価格倍率
    assert len(res.json()) == 19


def test_real_master_excel_excludes_struck_through_rows(client):
    """コード19957/19958/19960はコード・品名の双方に取り消し線があり、除外されること。"""
    for code in ("19957", "19958", "19960"):
        res = client.get("/api/master-items", params={"q": code})
        assert res.json() == [], f"code={code} should be excluded (strike-through)"


def test_real_master_excel_excludes_non_allowed_and_null_categories(client):
    """13品名以外のcategory (文章形式の特殊行4件・品名空欄1件) は取り込まれないこと。"""
    res = client.get("/api/master-items")
    items = res.json()
    categories = {item["category"] for item in items}
    assert categories == set(ALLOWED_CATEGORIES)
    assert None not in categories


def test_real_master_excel_tab_order_matches_business_order(client):
    """`GET /api/master-items` (全件) の返却順が、指定業務順で品名がグループ化されること。

    Frontendはこの順序をそのままタブ順として使う (品名一覧の二重管理を避けるため)。
    """
    res = client.get("/api/master-items")
    items = res.json()
    seen_order: list[str] = []
    for item in items:
        c = item["category"]
        if c not in seen_order:
            seen_order.append(c)
    assert seen_order == list(ALLOWED_CATEGORIES)


def test_real_master_excel_10_column_mapping(client):
    res = client.get("/api/master-items", params={"q": "11001"})
    item = res.json()[0]
    assert item["code"] == "11001"
    assert item["model"] == "OS2- 816"
    assert item["rating"] == "2.3*0.8*1.6"
    assert item["total_price_a"] == 315300
    assert item["box_parts_price"] == 61600
    assert item["painting_price"] == 89100
    assert item["setup_a"] == 216
    assert item["sheet_metal_price"] == 1096
    assert item["assembly_price"] == 351
    assert item["inspection_price"] == 15


def test_master_excel_path_and_sheet_are_the_configured_ones():
    # 実装が参照している設定値そのものを確認 (テストがこっそり別ファイルを見ないように)。
    assert MASTER_EXCEL_PATH.name == "estimate_master_a.xlsx"
    assert MASTER_EXCEL_SHEET == "Sheet2"
    assert MASTER_EXCEL_PATH.exists()


# --- 合成Excelを使ったUPSERT(再取込)・取り消し線判定・除外・異常系のテスト ---


def _write_master_excel(path, rows, strike_cells: set[tuple[int, int]] | None = None):
    """rows: 各行は [コード, 品名, 型式, 定格, 総合価格A, ...] のリスト。

    strike_cells: {(0-based row index, 0-based column index)} の集合。指定されたセルへ
    取り消し線フォント (`Font(strike=True)`) を設定する (列0=コード, 列1=品名)。
    文字列の中身ではなく、実際のセル書式で取り消し線を表現する (追加指示5章)。
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet2"
    ws.append(["コード", "品名", "型式", "定格", "総合価格A", "箱・部品価格",
               "塗装価格", "設A", "板金", "組立", "検査"])
    for row in rows:
        ws.append(row)
    for row_idx, col_idx in (strike_cells or set()):
        cell = ws.cell(row=row_idx + 2, column=col_idx + 1)  # +2: 1行目はヘッダ、1-indexed
        cell.font = Font(strike=True)
    wb.save(path)


@pytest.fixture()
def fresh_conn(tmp_path):
    db_path = tmp_path / "importer_test.db"
    run_migrations(db_path)
    with get_connection(db_path) as conn:
        yield conn


def test_import_upserts_new_and_existing_codes_preserving_id(tmp_path, fresh_conn):
    excel_path = tmp_path / "master_v1.xlsx"
    _write_master_excel(excel_path, [
        ["11001", ALLOWED_CATEGORIES[0], "OS2-816", "2.3*0.8*1.6", 100, 10, 20, 1, 2, 3, 4],
        ["18101", ALLOWED_CATEGORIES[4], "なし", "W+D=1500", -100, -100, 0, None, None, None, None],
    ])
    result1 = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result1.imported == 2
    assert result1.inserted == 2
    assert result1.updated == 0

    row = fresh_conn.execute(
        "SELECT id FROM estimate_master_items WHERE code = ?", ("11001",)
    ).fetchone()
    original_id = row["id"]

    # 再取込: 11001の値を更新 + 新規コード追加
    excel_path_v2 = tmp_path / "master_v2.xlsx"
    _write_master_excel(excel_path_v2, [
        ["11001", ALLOWED_CATEGORIES[0], "OS2-816", "2.3*0.8*1.6", 999, 10, 20, 1, 2, 3, 4],
        ["18101", ALLOWED_CATEGORIES[4], "なし", "W+D=1500", -100, -100, 0, None, None, None, None],
        ["99999", ALLOWED_CATEGORIES[6], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
    ])
    result2 = import_master_excel(fresh_conn, excel_path=excel_path_v2, sheet_name="Sheet2")
    assert result2.inserted == 1  # 99999のみ新規
    assert result2.updated == 2  # 11001, 18101 は既存コード

    row_after = fresh_conn.execute(
        "SELECT id, total_price_a FROM estimate_master_items WHERE code = ?", ("11001",)
    ).fetchone()
    # idが維持されること (Manual BBoxのmaster_item_id参照が壊れないため)
    assert row_after["id"] == original_id
    assert row_after["total_price_a"] == 999


def test_import_keeps_missing_values_as_null(tmp_path, fresh_conn):
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(excel_path, [
        ["18101", ALLOWED_CATEGORIES[4], "なし", "W+D=1500", -100, -100, 0, None, None, None, None],
    ])
    import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    row = fresh_conn.execute(
        "SELECT setup_a, sheet_metal_price FROM estimate_master_items WHERE code = ?", ("18101",)
    ).fetchone()
    assert row["setup_a"] is None
    assert row["sheet_metal_price"] is None


def test_import_skips_rows_without_code(tmp_path, fresh_conn):
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(excel_path, [
        [None, ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
        ["11001", ALLOWED_CATEGORIES[0], "OS2-816", "2.3*0.8*1.6", 100, 10, 20, 1, 2, 3, 4],
    ])
    result = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result.skipped_no_code == 1
    assert result.imported == 1


def test_import_raises_for_missing_file(fresh_conn, tmp_path):
    with pytest.raises(MasterImportError):
        import_master_excel(fresh_conn, excel_path=tmp_path / "does_not_exist.xlsx")


def test_import_raises_for_missing_sheet(tmp_path, fresh_conn):
    excel_path = tmp_path / "master.xlsx"
    wb = openpyxl.Workbook()
    wb.active.title = "WrongSheetName"
    wb.save(excel_path)
    with pytest.raises(MasterImportError):
        import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")


# --- 取り消し線判定 (追加指示4章/5章) ---


def test_import_excludes_row_with_struck_through_code(tmp_path, fresh_conn):
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(
        excel_path,
        [
            ["11001", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
            ["11002", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
        ],
        strike_cells={(0, 0)},  # 1行目のコードセルに取り消し線
    )
    result = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result.excluded_by_strike == 1
    assert result.excluded_by_strike_codes == ["11001"]
    assert result.imported == 1
    codes = {r["code"] for r in fresh_conn.execute("SELECT code FROM estimate_master_items")}
    assert codes == {"11002"}


def test_import_excludes_row_with_struck_through_category(tmp_path, fresh_conn):
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(
        excel_path,
        [
            ["11001", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
            ["11002", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
        ],
        strike_cells={(0, 1)},  # 1行目の品名セルに取り消し線
    )
    result = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result.excluded_by_strike == 1
    assert result.excluded_by_strike_codes == ["11001"]
    assert result.imported == 1


def test_import_excludes_row_with_both_code_and_category_struck_through(tmp_path, fresh_conn):
    """コード・品名の両方に取り消し線があっても二重に数えず1行として除外すること。"""
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(
        excel_path,
        [["11001", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7]],
        strike_cells={(0, 0), (0, 1)},
    )
    result = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result.excluded_by_strike == 1
    assert result.imported == 0


def test_import_imports_row_without_strike_normally(tmp_path, fresh_conn):
    """取り消し線が全く無い正常行は問題なく取り込まれること (回帰確認)。"""
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(
        excel_path,
        [["11001", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7]],
    )
    result = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result.excluded_by_strike == 0
    assert result.imported == 1


# --- 対象13品名への絞り込み (追加指示1章/2章) ---


def test_import_excludes_non_allowed_category(tmp_path, fresh_conn):
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(excel_path, [
        ["11001", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
        ["88888", "対象外の品名", "X", "Y", 1, 2, 3, 4, 5, 6, 7],
    ])
    result = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result.excluded_by_category == 1
    assert result.imported == 1
    codes = {r["code"] for r in fresh_conn.execute("SELECT code FROM estimate_master_items")}
    assert codes == {"11001"}


def test_import_excludes_null_category(tmp_path, fresh_conn):
    excel_path = tmp_path / "master.xlsx"
    _write_master_excel(excel_path, [
        ["11001", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
        ["77777", None, "X", "Y", 1, 2, 3, 4, 5, 6, 7],
    ])
    result = import_master_excel(fresh_conn, excel_path=excel_path, sheet_name="Sheet2")
    assert result.excluded_by_category == 1
    assert result.imported == 1
    codes = {r["code"] for r in fresh_conn.execute("SELECT code FROM estimate_master_items")}
    assert codes == {"11001"}


# --- 既存Masterの安全な整理 (追加指示8章/9章) ---


def _insert_manual_detection_referencing(conn: sqlite3.Connection, master_item_id: int) -> int:
    """テスト用に、指定Master ItemをManual BBoxとして参照するDetectionを1件作る。"""
    conn.execute("INSERT INTO drawing_files (original_filename) VALUES ('dummy.pdf')")
    file_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    conn.execute(
        """
        INSERT INTO drawing_pages
            (drawing_file_id, page_no, drawing_type, drawing_name, page_width, page_height)
        VALUES (?, 1, 'test', 'testpage', 1000, 800)
        """,
        (file_id,),
    )
    page_id = conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    conn.execute(
        """
        INSERT INTO detections
            (drawing_page_id, class_name, bbox_x, bbox_y, bbox_w, bbox_h, status,
             source_type, master_item_id)
        VALUES (?, 'manual-test', 0.1, 0.1, 0.05, 0.05, 'reviewed', 'manual', ?)
        """,
        (page_id, master_item_id),
    )
    return conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]


def test_reimport_removes_stale_master_item_not_referenced(tmp_path, fresh_conn):
    """今回の条件で無効になり、どのDetectionからも参照されていない行は削除される。"""
    excel_v1 = tmp_path / "master_v1.xlsx"
    _write_master_excel(excel_v1, [
        ["55555", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
    ])
    import_master_excel(fresh_conn, excel_path=excel_v1, sheet_name="Sheet2")
    assert fresh_conn.execute(
        "SELECT 1 FROM estimate_master_items WHERE code = ?", ("55555",)
    ).fetchone() is not None

    # v2では55555が対象外品名になった (=Excel側でも除外/変更されたと想定)
    excel_v2 = tmp_path / "master_v2.xlsx"
    _write_master_excel(excel_v2, [
        ["55555", "対象外の品名", "X", "Y", 1, 2, 3, 4, 5, 6, 7],
    ])
    result2 = import_master_excel(fresh_conn, excel_path=excel_v2, sheet_name="Sheet2")
    assert result2.removed_stale == 1
    assert result2.retained_invalid_referenced == []
    assert fresh_conn.execute(
        "SELECT 1 FROM estimate_master_items WHERE code = ?", ("55555",)
    ).fetchone() is None


def test_reimport_retains_master_item_referenced_by_manual_detection(tmp_path, fresh_conn):
    """Manual BBoxが参照しているMaster行は、無効化条件に該当しても削除しない (要件9)。"""
    excel_v1 = tmp_path / "master_v1.xlsx"
    _write_master_excel(excel_v1, [
        ["66666", ALLOWED_CATEGORIES[0], "X", "Y", 1, 2, 3, 4, 5, 6, 7],
    ])
    import_master_excel(fresh_conn, excel_path=excel_v1, sheet_name="Sheet2")
    master_id = fresh_conn.execute(
        "SELECT id FROM estimate_master_items WHERE code = ?", ("66666",)
    ).fetchone()["id"]
    _insert_manual_detection_referencing(fresh_conn, master_id)

    excel_v2 = tmp_path / "master_v2.xlsx"
    _write_master_excel(excel_v2, [
        ["66666", "対象外の品名", "X", "Y", 1, 2, 3, 4, 5, 6, 7],
    ])
    result2 = import_master_excel(fresh_conn, excel_path=excel_v2, sheet_name="Sheet2")

    assert result2.removed_stale == 0
    assert result2.retained_invalid_referenced == [(master_id, "66666")]
    # 削除されず残っていること (FK違反も起きない)
    still_there = fresh_conn.execute(
        "SELECT id FROM estimate_master_items WHERE code = ?", ("66666",)
    ).fetchone()
    assert still_there is not None
    assert still_there["id"] == master_id
    # Detection側の参照も無事に残っていること
    detection_row = fresh_conn.execute(
        "SELECT master_item_id FROM detections WHERE master_item_id = ?", (master_id,)
    ).fetchone()
    assert detection_row is not None
